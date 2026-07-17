# Bioindustrial-Park: BioSTEAM's Premier Biorefinery Models and Results
# Copyright (C) 2026-, Azhar Razin,
#                      Yalin Li <mailto.yalin.li@gmail.com>
#
# This module is under the UIUC open-source license. See
# github.com/BioSTEAMDevelopmentGroup/biosteam/blob/master/LICENSE.txt
# for license details.

"""
export_ad_report.py --> for Hauke

Updated exporter:
- includes raw biogas totals and composition in stream tables
- includes soil amendment and liquid digestate constituent breakdowns
- uses current model outputs case-by-case
"""

import argparse
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from biorefineries.sabre._chemicals import set_thermo
from biorefineries.sabre.utils import load_assumptions
from biorefineries.sabre.systems import create_ad_biomethane_system

ANNUAL_DAYS = 330.0
ANNUAL_HOURS = ANNUAL_DAYS * 24.0

CASES = [
    ("press_mill_only", "Press + Mill Only"),
    ("enzymatic", "Enzymatic"),
    ("peroxide", "Peroxide"),
    ("combined_PE", "Combined PE"),
    ("combined_PTE", "Combined PTE"),
]

UNIT_ORDER = ["PR", "PC", "EV", "ML", "PX", "HT", "EZ", "AD", "H2SR", "UP", "SP"]
DIGESTATE_COMPONENTS = [
    "Water", "Ash", "Glucan", "Xylan", "Mannan", "Galactan", "Arabinan",
    "Alginate", "Fucoidan", "Mannitol", "Protein", "OtherSolids", "Lignin",
]

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER_BOTTOM = Border(bottom=Side(style="medium", color="808080"))

def safe_sum_dict_values(d: Any) -> float:
    if not d:
        return 0.0
    if isinstance(d, dict):
        return float(sum(float(v) for v in d.values()))
    return 0.0

def annualize_add_opex(unit) -> float:
    add_opex = getattr(unit, "add_OPEX", None)
    if not add_opex:
        return 0.0
    if isinstance(add_opex, dict):
        return sum(float(v) for v in add_opex.values()) * ANNUAL_HOURS
    try:
        return float(add_opex) * ANNUAL_HOURS
    except Exception:
        return 0.0

def annualize_utility(unit) -> float:
    try:
        return float(unit.utility_cost) * ANNUAL_HOURS
    except Exception:
        return 0.0

def installed_cost(unit) -> float:
    for attr in ("installed_cost",):
        try:
            value = float(getattr(unit, attr))
            if math.isfinite(value):
                return value
        except Exception:
            pass
    for attr in ("installed_costs", "purchase_costs", "baseline_purchase_costs"):
        try:
            d = getattr(unit, attr)
            if d:
                value = safe_sum_dict_values(d)
                if math.isfinite(value):
                    return value
        except Exception:
            pass
    return 0.0

def unit_label(unit_id: str, unit) -> str:
    try:
        costs = getattr(unit, "baseline_purchase_costs", {})
        if costs:
            return " / ".join(str(k) for k in costs.keys())
    except Exception:
        pass
    return type(unit).__name__

def get_unit_map(system) -> dict[str, Any]:
    return {u.ID: u for u in system.units}

def find_biostimulant_stream(unit_map: dict[str, Any]):
    if "EV" in unit_map:
        return unit_map["EV"].outs[0]
    if "PC" in unit_map:
        return unit_map["PC"].outs[0]
    return None

def stream_total_mass(stream) -> float:
    return float(stream.F_mass)

def stream_component_mass(stream, cid: str) -> float:
    try:
        return float(stream.imass[cid])
    except Exception:
        return 0.0

def stream_component_kmol(stream, cid: str) -> float:
    try:
        return float(stream.imol[cid])
    except Exception:
        return 0.0

def moisture_pct(stream) -> float | None:
    total = float(stream.F_mass)
    if total <= 0:
        return None
    try:
        water = float(stream.imass["Water"])
    except Exception:
        water = 0.0
    return 100.0 * water / total

def dry_matter_pct(stream) -> float | None:
    total = float(stream.F_mass)
    if total <= 0:
        return None
    try:
        water = float(stream.imass["Water"])
    except Exception:
        water = 0.0
    return 100.0 * (total - water) / total

def gas_molpct(stream, cid: str) -> float | None:
    total = float(stream.F_mol)
    if total <= 0:
        return None
    return 100.0 * stream_component_kmol(stream, cid) / total

def style_title(ws, row: int, text: str, end_col: int = 7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left")
    return c

def style_section(ws, row: int, text: str, end_col: int = 6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True)
    c.fill = SECTION_FILL
    return c

def style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        c = ws.cell(row, col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = Border(bottom=THIN)
        c.alignment = Alignment(horizontal="center")

def set_widths(ws, widths: dict[int, float]):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

def build_case(quality: str, case_key: str):
    sys = create_ad_biomethane_system(quality=quality, pretreatment_case=case_key)
    sys.simulate()
    return sys, get_unit_map(sys)

def write_capex_opex_sheet(wb: Workbook, cases_data: dict[str, dict[str, Any]]):
    ws = wb.active
    ws.title = "CAPEX & OPEX"

    style_title(ws, 1, "AD Pathway — Unit CAPEX and OPEX by Pretreatment Case", end_col=6)
    ws.cell(2, 1, "Generated directly from BioSTEAM model | 330 days/yr")

    row = 4
    for case_key, case_label in CASES:
        case = cases_data[case_key]
        unit_map = case["unit_map"]

        style_section(ws, row, case_label, end_col=6)
        row += 1

        headers = ["ID", "Unit Name", "Installed Cost ($)", "add_OPEX ($/yr)", "Utility ($/yr)", "Total OPEX ($/yr)"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row, i, h)
        style_header_row(ws, row, len(headers))
        row += 1

        start_data = row
        for uid in UNIT_ORDER:
            unit = unit_map.get(uid)
            if unit is None:
                continue

            add_opex = annualize_add_opex(unit)
            utility = annualize_utility(unit)
            total_opex = add_opex + utility

            ws.cell(row, 1, uid)
            ws.cell(row, 2, unit_label(uid, unit))
            ws.cell(row, 3, installed_cost(unit))
            ws.cell(row, 4, add_opex)
            ws.cell(row, 5, utility)
            ws.cell(row, 6, total_opex)
            row += 1

        ws.cell(row, 1, "TOTAL")
        ws.cell(row, 3, f"=SUM(C{start_data}:C{row-1})")
        ws.cell(row, 4, f"=SUM(D{start_data}:D{row-1})")
        ws.cell(row, 5, f"=SUM(E{start_data}:E{row-1})")
        ws.cell(row, 6, f"=SUM(F{start_data}:F{row-1})")
        for c in range(1, 7):
            ws.cell(row, c).font = Font(bold=True)
            ws.cell(row, c).border = BORDER_BOTTOM
        row += 2

    set_widths(ws, {1: 10, 2: 36, 3: 18, 4: 18, 5: 18, 6: 18})
    for row_cells in ws.iter_rows(min_row=6, min_col=3, max_col=6):
        for c in row_cells:
            c.number_format = '$#,##0'

def write_stream_sheet(wb: Workbook, case_key: str, case_label: str, case_data: dict[str, Any]):
    title = "Stream Table" if case_key == "press_mill_only" else f"Stream Table — {case_label}"
    ws = wb.create_sheet(title)

    unit_map = case_data["unit_map"]

    pr = unit_map["PR"]
    ad = unit_map["AD"]
    h2sr = unit_map["H2SR"]
    up = unit_map["UP"]
    sp = unit_map["SP"]
    biostim = find_biostimulant_stream(unit_map)

    feed = pr.ins[0]
    ad_in = ad.ins[0]
    raw_biogas = ad.outs[0]
    treated_biogas = h2sr.outs[0]
    biomethane = up.outs[0]
    soil = sp.outs[0]
    liquid = sp.outs[1]

    style_title(ws, 1, f"AD Pathway — Stream Table ({case_label})", end_col=7)
    ws.cell(2, 1, f"{round(feed.F_mass, 0):,.0f} kg/hr wet feed | generated from model | 330 days/yr")

    headers = ["Stream", "Component", "Flow (kg/hr)", "Moisture (%)", "Dry Matter (%)", "Aux", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(4, i, h)
    style_header_row(ws, 4, len(headers))

    row = 5

    def add_row(stream_name, component, flow=None, moisture=None, drymatter=None, aux=None, note=None):
        nonlocal row
        ws.cell(row, 1, stream_name)
        ws.cell(row, 2, component)
        if flow is not None:
            ws.cell(row, 3, flow)
        if moisture is not None:
            ws.cell(row, 4, moisture)
        if drymatter is not None:
            ws.cell(row, 5, drymatter)
        if aux is not None:
            ws.cell(row, 6, aux)
        if note is not None:
            ws.cell(row, 7, note)
        row += 1

    add_row("Feed In", "Total wet", stream_total_mass(feed), moisture_pct(feed), dry_matter_pct(feed), note="Model feed stream")
    for cid in ["Water", "Ash", "Glucan", "Xylan", "Mannan", "Galactan", "Arabinan", "Alginate", "Fucoidan", "Mannitol", "Protein", "OtherSolids", "Lignin"]:
        mass = stream_component_mass(feed, cid)
        if mass > 0:
            add_row("", cid, mass)

    add_row("AD Inlet", "Total", stream_total_mass(ad_in), note="After press / mill / pretreatment")
    add_row("", "Biodegradable pool", ad.design_results.get("Biodegradable pool (kg/hr)"))
    add_row("", "Methane yield", aux=ad.ch4_kg_per_kg_vs_fed, note="kg CH4/kg VS fed")

    add_row("Raw Biogas", "Total", stream_total_mass(raw_biogas), aux=ad.design_results.get("Raw biogas total (Nm3/hr)"), note="kg/hr | Nm3/hr")
    for cid, label in [("Methane", "CH4"), ("CarbonDioxide", "CO2"), ("HydrogenSulfide", "H2S")]:
        mass = stream_component_mass(raw_biogas, cid)
        if mass > 0:
            pct = gas_molpct(raw_biogas, cid)
            add_row("", label, mass, aux=pct, note="mol%")

    add_row("Treated Biogas", "Total", stream_total_mass(treated_biogas), note="After H2S removal")
    h2s_ppm = h2sr.design_results.get("H2S outlet (ppm mol)")
    if h2s_ppm is not None:
        add_row("", "H2S outlet", aux=h2s_ppm, note="ppm mol")

    add_row("Biomethane Out", "Total", stream_total_mass(biomethane), note="Upgrader product")
    for cid, label in [("Methane", "CH4"), ("CarbonDioxide", "CO2 (residual)")]:
        mass = stream_component_mass(biomethane, cid)
        if mass > 0:
            pct = gas_molpct(biomethane, cid)
            add_row("", label, mass, aux=pct, note="mol%")

    if biostim is not None:
        add_row("Biostimulant", "Total concentrate", stream_total_mass(biostim), moisture_pct(biostim), dry_matter_pct(biostim), note="Pressate side-stream product")
        for cid in ["Water", "Alginate", "Fucoidan", "Mannitol", "Protein", "OtherSolids"]:
            mass = stream_component_mass(biostim, cid)
            if mass > 0:
                add_row("", cid, mass)

    add_row("Soil Amendment", "Total", stream_total_mass(soil), moisture_pct(soil), dry_matter_pct(soil), note="Digestate cake")
    for cid in DIGESTATE_COMPONENTS:
        mass = stream_component_mass(soil, cid)
        if mass > 0:
            add_row("", cid, mass)

    add_row("Liquid Digestate", "Total", stream_total_mass(liquid), moisture_pct(liquid), dry_matter_pct(liquid), note="Digestate centrate / filtrate")
    for cid in DIGESTATE_COMPONENTS:
        mass = stream_component_mass(liquid, cid)
        if mass > 0:
            add_row("", cid, mass)

    ws.cell(row + 1, 1, "Source: generated from BioSTEAM model output.")
    set_widths(ws, {1: 20, 2: 22, 3: 16, 4: 13, 5: 13, 6: 14, 7: 38})

    for r in ws.iter_rows(min_row=5, max_row=row, min_col=3, max_col=6):
        for c in r:
            if c.column in (4, 5, 6):
                c.number_format = '0.00'
            else:
                c.number_format = '#,##0'

def write_diagnostics_sheet(wb: Workbook, cases_data: dict[str, dict[str, Any]]):
    ws = wb.create_sheet("Diagnostics")
    style_title(ws, 1, "Model Export Diagnostics", end_col=5)
    headers = ["Severity", "Case", "Check", "Result", "Comment"]
    for i, h in enumerate(headers, start=1):
        ws.cell(3, i, h)
    style_header_row(ws, 3, len(headers))

    baseline = cases_data["press_mill_only"]
    b_sp = baseline["unit_map"]["SP"]
    b_up = baseline["unit_map"]["UP"]
    b_soil = stream_total_mass(b_sp.outs[0])
    b_liquid = stream_total_mass(b_sp.outs[1])
    b_bm = stream_total_mass(b_up.outs[0])

    row = 4
    for case_key, case_label in CASES:
        case = cases_data[case_key]
        unit_map = case["unit_map"]

        up = unit_map["UP"]
        sp = unit_map["SP"]
        soil = stream_total_mass(sp.outs[0])
        liquid = stream_total_mass(sp.outs[1])
        bm = stream_total_mass(up.outs[0])

        bm_total = stream_total_mass(up.outs[0])
        bm_comp = stream_component_mass(up.outs[0], "Methane") + stream_component_mass(up.outs[0], "CarbonDioxide")
        diff = bm_total - bm_comp
        severity = "WARN" if abs(diff) > 1.0 else "OK"
        ws.cell(row, 1, severity)
        ws.cell(row, 2, case_label)
        ws.cell(row, 3, "Biomethane total vs CH4+CO2")
        ws.cell(row, 4, f"Δ = {diff:,.1f} kg/h")
        ws.cell(row, 5, "Should be near zero unless other gas species are intentionally retained.")
        if severity == "WARN":
            for c in range(1, 6):
                ws.cell(row, c).fill = WARN_FILL
        row += 1

        gas_ratio = bm / b_bm if b_bm > 0 else float("nan")
        same_soil = abs(soil - b_soil) <= 1.0
        same_liquid = abs(liquid - b_liquid) <= 1.0
        severity = "WARN" if (case_key != "press_mill_only" and gas_ratio > 1.2 and same_soil and same_liquid) else "OK"
        ws.cell(row, 1, severity)
        ws.cell(row, 2, case_label)
        ws.cell(row, 3, "Gas increases but digestate unchanged")
        ws.cell(row, 4, f"biomethane ratio = {gas_ratio:.2f}x; same digestate = {same_soil and same_liquid}")
        ws.cell(row, 5, "AD water-balance / digestate reporting may still need review.")
        if severity == "WARN":
            for c in range(1, 6):
                ws.cell(row, c).fill = WARN_FILL
        row += 1

        ht = unit_map.get("HT")
        if ht is not None:
            util = annualize_utility(ht)
            severity = "WARN" if util <= 0.0 else "OK"
            ws.cell(row, 1, severity)
            ws.cell(row, 2, case_label)
            ws.cell(row, 3, "Thermal pretreatment annual utility")
            ws.cell(row, 4, f"${util:,.0f}/yr")
            ws.cell(row, 5, "If zero, check heat utility accounting or reporting.")
            if severity == "WARN":
                for c in range(1, 6):
                    ws.cell(row, c).fill = WARN_FILL
            row += 1

    set_widths(ws, {1: 12, 2: 18, 3: 30, 4: 26, 5: 60})

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--quality", default="pelagic_high_quality")
    parser.add_argument(
        "--out",
        default=str(SCRIPT_DIR.parent / "results" / "SaBRe_Unit_CAPEX_OPEX_generated.xlsx"),
    )
    args = parser.parse_args()

    set_thermo()
    load_assumptions()

    cases_data: dict[str, dict[str, Any]] = {}
    for case_key, case_label in CASES:
        system, unit_map = build_case(args.quality, case_key)
        cases_data[case_key] = {
            "label": case_label,
            "system": system,
            "unit_map": unit_map,
        }

    wb = Workbook()
    write_capex_opex_sheet(wb, cases_data)
    for case_key, case_label in CASES:
        write_stream_sheet(wb, case_key, case_label, cases_data[case_key])
    write_diagnostics_sheet(wb, cases_data)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved workbook to {out_path}")

if __name__ == "__main__":
    main()
