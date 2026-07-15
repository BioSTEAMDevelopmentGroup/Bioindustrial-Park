# Bioindustrial-Park: BioSTEAM's Premier Biorefinery Models and Results
# Copyright (C) 2026-, Azhar Razin,
#                      Yalin Li <mailto.yalin.li@gmail.com>
#
# This module is under the UIUC open-source license. See
# github.com/BioSTEAMDevelopmentGroup/biosteam/blob/master/LICENSE.txt
# for license details.

"""
generate_ad_stream_table.py
============================
Runs all five AD pretreatment cases at near-zero feed cost and generates
a BioSTEAM stream output Excel table matching Hauke's format:
  one row per stream per case, blank row between cases.

Output: AD_stream_table.xlsx
"""

import sys
from pathlib import Path

import biosteam as bst
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from biorefineries.sabre._chemicals import set_thermo
from biorefineries.sabre.systems import create_ad_biogas_system

OUT = SCRIPT_DIR.parent / "results"
OUT.mkdir(parents=True, exist_ok=True)

# ── Configuration ──────────────────────────────────────────────────────────

PRETREATMENT_CASES = [
    "press_mill_only",
    "enzymatic",
    "peroxide",
    "combined_PE",
    "combined_PTE",
]

FEED_PRICE = 0.00  # near-zero baseline

# Streams to capture (in order)
# Try multiple possible IDs for streams that vary by system version
STREAM_IDS = [
    "sargassum_feed",
    "pressed_cake",
    "pressate",
    "biostimulant_membrane_concentrate",
    "pressate_permeate",
    "milled_biomass",
    "milling_losses",
    # pretreatment intermediates (only present for relevant cases)
    "peroxide_treated_biomass",
    "enzyme_treated_biomass",
    "heated_biomass",
    "combined_PE_treated_biomass",
    "combined_PTE_treated_biomass",
    # biogas streams - try both naming conventions
    "raw_biogas",
    "biogas_raw",
    "treated_biogas",
    "biomethane",
    "offgas_co2",
    "spent_h2s_media",
    "digestate",
    "soil_amendment",
    "liquid_digestate",
]

# Components to report (kg/hr columns)
COMPONENTS = [
    "Water",
    "Methane",
    "CarbonDioxide",
    "HydrogenSulfide",
    "Glucan",
    "Alginate",
    "Fucoidan",
    "Mannitol",
    "Protein",
    "OtherSolids",
    "Ash",
]

# ── Helper functions ───────────────────────────────────────────────────────

def _safe_imass(stream, cid):
    try:
        return float(stream.imass[cid])
    except Exception:
        return 0.0


def _stream_row(case_name, stream_id, stream):
    """Extract one row of data from a stream."""
    chem_ids = set(stream.chemicals.IDs)

    water_kg = _safe_imass(stream, "Water") if "Water" in chem_ids else 0.0
    total_kg  = float(stream.F_mass)
    dry_kg    = max(total_kg - water_kg, 0.0)

    # VS = total - Water - Ash
    ash_kg = _safe_imass(stream, "Ash") if "Ash" in chem_ids else 0.0
    vs_kg  = max(dry_kg - ash_kg, 0.0)

    row = {
        "case":          case_name,
        "stream":        stream_id,
        "phase":         stream.phase if len(stream.phase) == 1 else "mixed",
        "F_mass_kg_hr":  total_kg,
        "water_wt_frac": water_kg / total_kg if total_kg > 0 else 0.0,
        "TS_kg_hr":      dry_kg,
        "VS_kg_hr":      vs_kg,
    }

    for cid in COMPONENTS:
        col = f"{cid}_kg_hr"
        row[col] = _safe_imass(stream, cid) if cid in chem_ids else 0.0

    return row


# Canonical order for display — streams not in this list appear at the end
STREAM_ORDER = [
    "sargassum_feed",
    "pressed_cake", "pressate",
    "biostimulant_membrane_concentrate", "pressate_permeate",
    "milled_biomass", "milling_losses",
    "peroxide_treated_biomass", "heated_biomass",
    "enzyme_treated_biomass",
    "combined_PE_treated_biomass", "combined_PTE_treated_biomass",
    "raw_biogas", "biogas_raw",
    "spent_h2s_media", "treated_biogas",
    "biomethane", "offgas_co2",
    "digestate", "soil_amendment", "liquid_digestate",
]


def run_case(case):
    """Build, simulate, and extract stream rows for one pretreatment case."""
    bst.main_flowsheet.clear()
    set_thermo()

    sys = create_ad_biogas_system(
        quality="pelagic_high_quality",
        pretreatment_case=case,
    )
    sys.feeds[0].price = FEED_PRICE
    sys.simulate()

    # Collect ALL streams from the flowsheet — no hardcoded IDs needed
    fs = sys.flowsheet.stream
    all_streams = {s.ID: s for s in fs}

    # Sort by canonical order, then alphabetically for anything not listed
    def sort_key(sid):
        try:
            return (STREAM_ORDER.index(sid), sid)
        except ValueError:
            return (len(STREAM_ORDER), sid)

    rows = []
    for sid in sorted(all_streams.keys(), key=sort_key):
        s = all_streams[sid]
        if float(s.F_mass) > 0:  # skip empty streams
            rows.append(_stream_row(case, sid, s))

    return rows


# ── Build data ─────────────────────────────────────────────────────────────

all_rows = []
if __name__ == "__main__":
    for case in PRETREATMENT_CASES:
        print(f"  Running {case}...")
        case_rows = run_case(case)
        all_rows.extend(case_rows)
        all_rows.append({})  # blank separator row

    df = pd.DataFrame(all_rows)

    # ── Write Excel ────────────────────────────────────────────────────────────

    OUTFILE = OUT / "AD_stream_table.xlsx"

    # Column order
    meta_cols  = ["case", "stream", "phase", "F_mass_kg_hr",
                   "water_wt_frac", "TS_kg_hr", "VS_kg_hr"]
    comp_cols  = [f"{c}_kg_hr" for c in COMPONENTS]
    all_cols   = meta_cols + comp_cols

    # Ensure all columns exist
    for c in all_cols:
        if c not in df.columns:
            df[c] = 0.0

    df = df[all_cols]
    df.to_excel(OUTFILE, index=False, sheet_name="Streams")

    # ── Formatting ─────────────────────────────────────────────────────────────

    wb = load_workbook(OUTFILE)
    ws = wb["Streams"]

    # Styles
    HEADER_FILL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    CASE_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    CASE_FONT    = Font(name="Arial", bold=True, size=9, color="1F3864")
    BODY_FONT    = Font(name="Arial", size=9)
    BLANK_FILL   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT        = Alignment(horizontal="right",  vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(all_cols)

    # Header row
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        # Nicer header labels
        label = col_name.replace("_kg_hr", "\n(kg/hr)").replace("_", " ")
        cell.value = label

    # Data rows
    for row_idx in range(2, ws.max_row + 1):
        row_data = df.iloc[row_idx - 2]
        is_blank = pd.isna(row_data.get("case")) or row_data.get("case") == ""
        is_new_case = not is_blank and (
            row_idx == 2 or
            df.iloc[row_idx - 3].get("case") != row_data.get("case")
        )

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER

            if is_blank:
                cell.fill = BLANK_FILL
                cell.value = None
                continue

            col_name = all_cols[col_idx - 1]

            if is_new_case and col_idx == 1:
                cell.fill = CASE_FILL
                cell.font = CASE_FONT
            elif col_name in ("case", "stream", "phase"):
                cell.font = CASE_FONT if is_new_case else BODY_FONT
                cell.fill = CASE_FILL if is_new_case else PatternFill()
                cell.alignment = LEFT
            else:
                cell.font = BODY_FONT
                cell.alignment = RIGHT

            val = row_data.get(col_name)
            if pd.isna(val):
                cell.value = None
            elif col_name == "water_wt_frac":
                cell.value = round(float(val), 6)
                cell.number_format = "0.0000"
            elif col_name in ("case", "stream", "phase"):
                cell.value = str(val) if val else None
            else:
                cell.value = round(float(val), 2)
                cell.number_format = "#,##0.00"

    # Column widths
    COL_WIDTHS = {
        "case":          18,
        "stream":        28,
        "phase":          7,
        "F_mass_kg_hr":  14,
        "water_wt_frac": 12,
        "TS_kg_hr":      12,
        "VS_kg_hr":      12,
    }
    DEFAULT_WIDTH = 14

    for col_idx, col_name in enumerate(all_cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            COL_WIDTHS.get(col_name, DEFAULT_WIDTH)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Row height — header taller for wrapped text
    ws.row_dimensions[1].height = 32

    wb.save(OUTFILE)
    print(f"\nSaved: {OUTFILE}")